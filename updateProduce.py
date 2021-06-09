#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

"""
• Opens spreadsheet file - ProdceSales.xlsx
• For each row, checks whether the value in column A is Celery, Garlic, or Lemon.
• If it is, updates the price in column B.
• Saves the spreadsheet to a new file so that the old spreadsheet remains unchanged. 
"""

import openpyxl


wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

# The produce types and their updated prices.
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# Loop through the rows and update prices.
for rowNum in range(2, sheet.max_row):  # The loop skips the first row.
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save("UpdateProduceSales.xlsx")
